import pandas as pd
import numpy as np
"""

# Sample customer data
# customer_data = {
#     'CustomerID': np.arange(1, 6),
#     'Name': ['John', 'Alice', 'Bob', 'Eva', 'Mike'],
#     'Email': ['john@example.com', 'alice@example.com', 'bob@example.com', 'eva@example.com', 'mike@example.com'],
#     'Phone': ['123-456-7890', '987-654-3210', '456-789-0123', '789-012-3456', '012-345-6789'],
#     'TotalSpent': [50.0, 30.5, 75.2, 40.0, 60.8]
# }

# # Create a DataFrame from the customer data 
# df = pd.DataFrame(customer_data)
# print(df)

# # Export the DataFrame to Excel
# excel_file_path = 'bakery_customer_data.xlsx'
# # df.to_excel(excel_file_path, index=False)

# print(f"Customer data exported to {excel_file_path}")
customer_data = {
    'Name': ['aditya', 'abhishek', 'ketan', 'jeetu', 'name'],
    'Email': ['<EMAIL>', '<EMAIL>', '<EMAIL>', '<EMAIL>', '<EMAIL>'],
    'Phone': ['546464', '246546556', '6546554', '4454544', '3654365365'],
    'TotalSpent': ['4500', '5200', '7800', '5600', '8500']
}

customer_data_main = {
    # "customer_id_main": np.arange(len('Name[]')),
    "customer_name_main": customer_data['Name'],
    "customer_email_main": customer_data['Email'],
    "customer_phone_main": customer_data['Phone'],
    "customer_total_spent_main": customer_data['TotalSpent']

}

df_main = pd.DataFrame(customer_data_main)

print(df_main)
# print(f"Customer data exported to {excel_file_path}")
"""
import openpyxl

# Define bakery customer data as a list of dictionaries
customers = [
    {"Name": "Alice", "Email": "alice@example.com", "Phone": "9876543210"},
    {"Name": "Bob", "Email": "bob@example.com", "Phone": "1234567890"},
    {"Name": "Charlie", "Email": "charlie@example.com", "Phone": "0987654321"},
]

# Create a new Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Write headers to the first row
sheet.cell(row=1, column=1).value = "Name"
sheet.cell(row=1, column=2).value = "Email"
sheet.cell(row=1, column=3).value = "Phone"

# Write customer data to rows
for i, customer in enumerate(customers, start=2):
    sheet.cell(row=i, column=1).value = customer["Name"]
    sheet.cell(row=i, column=2).value = customer["Email"]
    sheet.cell(row=i, column=3).value = customer["Phone"]

# Save the Excel file
wb.save("bakery_customers.xlsx")

print("Bakery customer data exported to bakery_customers.xlsx")
